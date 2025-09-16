# app_dragdrop_excel_reports.py
# 変更点: (1) PDF関連コードの完全削除 (2) ボタン文言/タイトルUIの簡素化
# 強化: (A) ヘッダ検出で「払出数」を数量より優先 (B) LOT抽出で「Lot.」も許可
#       (C) 行キャリー＋集約ロジックに「シリアルのみぶら下がり＆Lot No.空」の特例対応
#       (D) 新しい型番ブロック開始時に last_lotno / last_exp_norm を確実にリセット（誤キャリー防止）
# ------------------------------------------------------------
# pip install streamlit pandas openpyxl requests
# 実行: streamlit run app_dragdrop_excel_reports.py
# ------------------------------------------------------------
import io, os, re, time, datetime as dt
from typing import List, Tuple, Optional, Dict, Any

import streamlit as st
import pandas as pd
import requests
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter

HEADERS = ["工程名","LOT","型番","Lot No.","払出数","有効期限","ファイル名"]

# ===================== ユーティリティ =====================
def norm(s) -> str:
    if s is None: return ""
    return str(s).replace("\r"," ").replace("\n"," ").replace("　"," ").strip()

def normalize_date(s: Optional[str]) -> Optional[str]:
    if not s: return None
    m = re.search(r"(\d{4})[./-](\d{1,2})[./-](\d{1,2})", s)
    if not m: return None
    y, mth, d = int(m.group(1)), int(m.group(2)), int(m.group(3))
    try:
        dtv = pd.Timestamp(year=y, month=mth, day=d)
        return f"{dtv.year}/{dtv.month}/{dtv.day}"
    except Exception:
        return None

def autosize(ws):
    for col in range(1, ws.max_column + 1):
        letter = get_column_letter(col)
        max_len = 0
        for cell in ws[letter]:
            val = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(val))
        ws.column_dimensions[letter].width = min(max_len + 2, 80)

# 数量の強化正規化：数値型/小数/カンマ/全角/単位付きもOK
def _to_int_qty(q) -> Optional[int]:
    if q is None or (isinstance(q, float) and pd.isna(q)):
        return None
    if isinstance(q, (int, float)):
        return int(round(float(q)))
    s = str(q).strip()
    s = s.translate(str.maketrans("０１２３４５６７８９－．，", "0123456789-.,")).replace(",", "")
    m = re.match(r"^\s*([+-]?\d+(?:\.\d+)?)", s)
    if not m:
        return None
    try:
        return int(round(float(m.group(1))))
    except Exception:
        return None

# NEW: ファイル名から「返庫」判定（Excelにのみ適用）
def is_henko_from_name(filename_wo_ext: str) -> bool:
    return "返庫" in (filename_wo_ext or "")

# ===================== Excelヘッダ検出（2段対応） =====================
HEADER_KEYS = {
    "model": ["型番","品目","品番","型 式"],
    "lotno": ["Lot No","LotNo","LOT NO","ロット","Lot"],
    # qty は「払出数系 ＞ 数量系」で優先
    "qty_hi": ["払出数","払い出し","払出","出庫","出数"],
    "qty_lo": ["数量","個数","数"],
    "exp":   ["有効期限","期限","賞味期限","Exp","有効期日"],
}

def _n(cell):
    if pd.isna(cell): return ""
    return str(cell).strip().replace("　","").replace("\n"," ").replace("\r"," ")

def detect_header(df: pd.DataFrame, scan_rows: int = 40) -> dict|None:
    def choose_col_by_priority(cells: list[str], high: list[str], low: list[str]) -> Optional[int]:
        lows = [c.lower() for c in cells]
        for kw in high:
            kwl = kw.lower()
            for c, txt in enumerate(lows):
                if kwl in txt: return c
        for kw in low:
            kwl = kw.lower()
            for c, txt in enumerate(lows):
                if kwl in txt: return c
        return None

    def first_hit(cells: list[str], keys: list[str]) -> Optional[int]:
        lows = [c.lower() for c in cells]
        for kw in keys:
            kwl = kw.lower()
            for c, txt in enumerate(lows):
                if kwl in txt: return c
        return None

    def hit_from_row(cells: list[str]) -> dict:
        hit={}
        lot_c = first_hit(cells, HEADER_KEYS["lotno"])
        if lot_c is not None: hit["lotno"] = lot_c
        exp_c = first_hit(cells, HEADER_KEYS["exp"])
        if exp_c is not None: hit["exp"] = exp_c
        model_c = first_hit(cells, HEADER_KEYS["model"])
        if model_c is not None: hit["model"] = model_c
        qty_c = choose_col_by_priority(cells, HEADER_KEYS["qty_hi"], HEADER_KEYS["qty_lo"])
        if qty_c is not None: hit["qty"] = qty_c
        if sum(1 for k in ["lotno","qty","exp"] if k in hit) >= 2 and "model" in hit:
            return hit
        return {}

    scan_rows = min(len(df), scan_rows)
    # 1段
    for r in range(scan_rows):
        row = [_n(x) for x in df.iloc[r,:].tolist()]
        if not any(row): continue
        hit = hit_from_row(row)
        if hit: return {"row": r, **hit}
    # 2段（上下マージ）
    for r in range(scan_rows-1):
        row1 = [_n(x) for x in df.iloc[r,:].tolist()]
        row2 = [_n(x) for x in df.iloc[r+1,:].tolist()]
        width = max(len(row1), len(row2))
        combo=[]
        for c in range(width):
            a = row1[c] if c < len(row1) else ""
            b = row2[c] if c < len(row2) else ""
            combo.append((a or b) if (a or b) else "")
        if not any(combo): continue
        hit = hit_from_row(combo)
        if hit: return {"row": r, **hit}
    return None

def extract_koutei_lot_from_sheet(df: pd.DataFrame, max_scan_rows:int=8, max_scan_cols:int=8) -> Tuple[Optional[str], Optional[str]]:
    koutei=None; lot=None
    rows=min(len(df),max_scan_rows); cols=min(df.shape[1],max_scan_cols)
    # 工程名（上部の最初の非空セル）
    for r in range(rows):
        for c in range(cols):
            v=_n(df.iat[r,c])
            if v and not re.search(r"\b(lot|ロット)\b", v, flags=re.I):
                koutei=v; break
        if koutei: break
    # LOT（Lot: / Lot. / ロット: を許可）
    lot_pat = re.compile(r"(?:\bLot\b\.?|ロット)\s*[：:\.\s]\s*([^\s]+)", re.I)
    for r in range(rows):
        for c in range(cols):
            s=_n(df.iat[r,c])
            if not s: continue
            m=lot_pat.search(s)
            if m: lot=m.group(1).strip(); break
        if lot: break
    return koutei, lot

# ===================== シート選択（編集用優先・日付優先・確認系除外） =====================
EXCLUDED_SHEET_PATTERNS = [
    r"基本シート", r"^確認", r"確認用", r"確認\s*\(編集後\)", r"確認\(編集後\)",
    r"チェック", r"Check", r"DL用"
]
DATE_SHEET_RE = re.compile(r"(?:^|\s)(\d{4})[./-](\d{1,2})[./-](\d{1,2})(?:\s|$)")

def choose_target_sheet(sheet_names: list[str]) -> tuple[str, str]:
    if "編集用" in sheet_names:
        return "編集用", "編集用が最優先"
    dated: list[tuple[dt.date, str]] = []
    for s in sheet_names:
        m = DATE_SHEET_RE.search(s)
        if m:
            y, mth, d = map(int, m.groups())
            try: dated.append((dt.date(y, mth, d), s))
            except: pass
    if dated:
        dated.sort(reverse=True)
        return dated[0][1], f"日付シート優先（最新={dated[0][0].isoformat()}）"
    for s in sheet_names:
        if any(re.search(pat, s) for pat in EXCLUDED_SHEET_PATTERNS):
            continue
        return s, "除外を除いた先頭シート"
    return sheet_names[0], "フォールバック（先頭）"

# ===================== Excel明細抽出（キャリー＋集約＋特例＋境界リセット） =====================
def parse_excel_table(
    df: pd.DataFrame,
    header_map: dict,
    koutei: str,
    lot: str,
    file_label: str,
    qty_sign:int=1,
    require_lotno: bool=True,
    require_exp: bool=True,
) -> tuple[list[dict], dict]:
    """
    - 「型番セルが1回のみ」「同一Lot No.が下に複数行（数量だけ1ずつ等）」をサポート。
    - 行走査時に last_model / last_lotno / last_exp_norm をキャリー。
    - (model, lotno, exp_norm) 単位で数量を集約（qty_sign適用後）→ 出力。
    - 特例: 「シリアルだけが下にぶら下がり、Lot No.欄が全体で空」のブロックは、
            Lot No.空のまま（許容）で型番行の払出数を1行にまとめて出力。
            （UIでLot No.必須=ONでもブロック内にLot No.が1つも無ければ許容）
    - 重要: 新しい“型番”を検知した時点で、last_lotno / last_exp_norm を必ず None にリセットし、
            前ブロックのLot/期限が誤ってキャリーされるのを防止。
    """
    start=header_map["row"]+1
    mc,lc,qc,ec = header_map["model"],header_map["lotno"],header_map["qty"],header_map["exp"]
    sub=df.iloc[start:].copy().reset_index(drop=True)

    stats = {"空行":0, "型番欠落":0, "LotNo欠落":0, "数量不正":0, "数量=0":0, "日付不正":0}
    agg: Dict[tuple, int] = {}

    last_model: Optional[str] = None
    last_lotno: Optional[str] = None
    last_exp_norm: Optional[str] = None

    def _cell(row, idx):
        return row.iloc[idx] if idx < len(row) else None

    def _str_or_none(v):
        s = None if pd.isna(v) or str(v).strip()=="" else str(v).strip()
        return s

    def has_any_lotno_until_next_model(start_i: int) -> bool:
        """現在の行以降、次のモデルが出るまでの間にLotNoが1つでもあるか"""
        for j in range(start_i+1, len(sub)):
            rj = sub.iloc[j,:]
            model_j = _str_or_none(_cell(rj, mc))
            lotno_j = _str_or_none(_cell(rj, lc))
            if model_j:  # 次のブロック開始
                break
            if lotno_j:
                return True
        return False

    def lookahead_first_exp(start_i: int) -> Optional[str]:
        """現在の行を含め、次のモデルが出るまでに最初に見つかった期限を返す"""
        for j in range(start_i, len(sub)):
            rj = sub.iloc[j,:]
            model_j = _str_or_none(_cell(rj, mc))
            if j > start_i and model_j:  # 次ブロックに入ったら終了
                break
            expv_j = _str_or_none(_cell(rj, ec))
            exp_norm_j = normalize_date(expv_j) if expv_j else None
            if exp_norm_j:
                return exp_norm_j
        return None

    for i in range(len(sub)):
        row=sub.iloc[i,:]

        model_raw=_cell(row, mc)
        lotno_raw=_cell(row, lc)
        qty_raw  =_cell(row, qc)
        exp_raw  =_cell(row, ec)

        model = _str_or_none(model_raw)
        lotno = _str_or_none(lotno_raw)
        qty_i = _to_int_qty(qty_raw)
        exp_s = _str_or_none(exp_raw)
        exp_norm = normalize_date(exp_s) if exp_s else None

        # 完全空行
        if not any([model, lotno, (qty_i is not None), (exp_s is not None and exp_s!="")]):
            stats["空行"] += 1
            continue

        # ★ ブロック境界検知：この行に model があれば新ブロック開始
        if model:
            last_model = model
            # 前ブロックのLot/期限はここで確実に捨てる（誤キャリー防止）
            last_lotno = None
            last_exp_norm = None

        # 入力がある項目だけ last_* を更新
        if lotno: last_lotno = lotno
        if exp_norm: last_exp_norm = exp_norm

        cur_model = last_model
        cur_lotno = last_lotno
        cur_exp   = last_exp_norm

        if not cur_model:
            stats["型番欠落"] += 1
            continue

        # 数量チェック
        if qty_i is None:
            # 期限だけやシリアル行などは既にキャリー済みなのでエラーにしない
            continue
        if qty_i == 0:
            stats["数量=0"] += 1
            continue
        qty_i = abs(qty_i) * qty_sign

        # 期限がこの時点で未確定なら、同ブロック内から先読み
        if require_exp and not cur_exp:
            peek_exp = lookahead_first_exp(i)
            if peek_exp:
                cur_exp = peek_exp
                last_exp_norm = peek_exp
            else:
                stats["日付不正"] += 1
                continue

        # Lot No.必須だが、ブロック内にLotNoが1つも無い=シリアルだけの特例は許容
        if require_lotno and not cur_lotno:
            if has_any_lotno_until_next_model(i):
                stats["LotNo欠落"] += 1
                continue
            else:
                cur_lotno = ""  # 特例：空のまま出力

        key = (cur_model, cur_lotno or "", cur_exp or "")
        agg[key] = agg.get(key, 0) + qty_i

    out: List[Dict[str, Any]] = []
    for (model, lotno, exp_norm), qty_sum in agg.items():
        out.append({
            "工程名": koutei or "",
            "LOT": lot or "",
            "型番": model,
            "Lot No.": lotno,
            "払出数": qty_sum,
            "有効期限": exp_norm or "",
            "ファイル名": file_label
        })

    return out, stats

# ===================== 集計（品名ごと／工程ごと） =====================
def ensure_sheet(wb, name, headers):
    ws = wb[name] if name in wb.sheetnames else wb.create_sheet(name)
    if ws.max_row < 1 or all(ws.cell(row=1,column=i+1).value is None for i in range(len(headers))):
        for i,h in enumerate(headers,1): ws.cell(row=1,column=i).value = h
    return ws

def clear_sheet_body(ws):
    if ws.max_row > 1:
        ws.delete_rows(idx=2, amount=ws.max_row-1)

def read_sheet_as_records(ws) -> list[dict]:
    headers = [str(ws.cell(row=1, column=c).value or "").strip() for c in range(1, ws.max_column+1)]
    recs=[]
    for r in range(2, ws.max_row+1):
        row={}; empty=True
        for c,h in enumerate(headers,1):
            v = ws.cell(row=r, column=c).value
            if v not in (None,""): empty=False
            row[h]=v
        if not empty: recs.append(row)
    return recs

def build_name_map_from_master(wb) -> dict:
    if "品名マスタ" not in wb.sheetnames:
        ws = wb.create_sheet("品名マスタ")
        ws.cell(row=1, column=1, value="品名")
        ws.cell(row=1, column=2, value="型番")
        return {}
    ws = wb["品名マスタ"]
    mp={}
    for r in range(2, ws.max_row+1):
        pname = ws.cell(row=r, column=1).value
        model = ws.cell(row=r, column=2).value
        if model:
            mp[str(model).strip()] = ("" if pname in (None,"") else str(pname).strip())
    return mp

def refresh_reports_in_workbook(wb, edit_sheet_name="編集用"):
    if edit_sheet_name not in wb.sheetnames:
        return
    ws_edit = wb[edit_sheet_name]
    records = read_sheet_as_records(ws_edit)
    ws_by_item = ensure_sheet(wb, "品名ごと", ["品名","型番","払出数合計"])
    ws_by_proc = ensure_sheet(wb, "工程ごと", ["工程名","品名","型番","払出数合計"])
    clear_sheet_body(ws_by_item); clear_sheet_body(ws_by_proc)
    if not records:
        autosize(ws_by_item); autosize(ws_by_proc); return
    for rec in records:
        try: rec["払出数"] = int(rec.get("払出数",0))
        except: rec["払出数"] = 0
    name_map = build_name_map_from_master(wb)
    sum_by_model={}
    for rec in records:
        model = str(rec.get("型番") or "").strip()
        if not model: continue
        sum_by_model[model] = sum_by_model.get(model,0) + rec["払出数"]
    for model in sorted(sum_by_model.keys()):
        pname = name_map.get(model, "")
        ws_by_item.append([pname, model, sum_by_model[model]])
    autosize(ws_by_item)
    sum_by_proc_model={}
    for rec in records:
        proc  = str(rec.get("工程名") or "").strip()
        model = str(rec.get("型番") or "").strip()
        if not proc or not model: continue
        key=(proc, model)
        sum_by_proc_model[key] = sum_by_proc_model.get(key,0) + rec["払出数"]
    for (proc, model) in sorted(sum_by_proc_model.keys(), key=lambda x:(x[0],x[1])):
        pname = name_map.get(model, "")
        ws_by_proc.append([proc, pname, model, sum_by_proc_model[(proc, model)]])
    autosize(ws_by_proc)

# ===================== “編集用”追記＋レポート再作成 =====================
def update_workbook_with_rows(base_xlsx_bytes: bytes|None, rows: List[Dict[str,Any]], sheet_name:str="編集用") -> bytes:
    if base_xlsx_bytes:
        wb = load_workbook(io.BytesIO(base_xlsx_bytes))
    else:
        wb = Workbook()
        wb.active.title = sheet_name
        ws0 = wb[sheet_name]; ws0.append(HEADERS)
        ws_m = wb.create_sheet("品名マスタ")
        ws_m.cell(row=1, column=1, value="品名")
        ws_m.cell(row=1, column=2, value="型番")
        ensure_sheet(wb, "品名ごと", ["品名","型番","払出数合計"])
        ensure_sheet(wb, "工程ごと", ["工程名","品名","型番","払出数合計"])

    ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.create_sheet(sheet_name)
    if ws.max_row < 1 or all(ws.cell(row=1,column=c).value is None for c in range(1,len(HEADERS)+1)):
        for c,h in enumerate(HEADERS,1): ws.cell(row=1,column=c).value = h

    for r in rows:
        q = _to_int_qty(r.get("払出数"))
        if q is None or q == 0:
            continue
        ws.append([r.get(h,"") for h in HEADERS])
    autosize(ws)
    refresh_reports_in_workbook(wb, edit_sheet_name=sheet_name)
    bio=io.BytesIO(); wb.save(bio); return bio.getvalue()

# ===================== Copilot Studio（Direct Line）連携テスト =====================
def copilot_directline_test(secret: str, test_message: str = "ping") -> tuple[bool, str]:
    try:
        headers = {"Authorization": f"Bearer {secret}", "Content-Type": "application/json"}
        r = requests.post("https://directline.botframework.com/v3/directline/conversations", headers=headers, json={})
        if r.status_code != 201:
            return False, f"start failed: {r.status_code} {r.text}"
        conv = r.json(); conv_id = conv.get("conversationId")
        if not conv_id:
            return False, "no conversationId"
        send_url = f"https://directline.botframework.com/v3/directline/conversations/{conv_id}/activities"
        payload = {"type":"message","from":{"id":"user1"}, "text": test_message}
        r2 = requests.post(send_url, headers=headers, json=payload)
        if r2.status_code not in (200, 201):
            return False, f"send failed: {r2.status_code} {r2.text}"
        get_url = f"https://directline.botframework.com/v3/directline/conversations/{conv_id}/activities"
        for _ in range(5):
            r3 = requests.get(get_url, headers=headers)
            if r3.status_code != 200:
                time.sleep(0.5); continue
            acts = r3.json().get("activities", [])
            bot_msgs = [a for a in acts if a.get("from",{}).get("id","").lower() != "user1" and a.get("type")=="message"]
            if bot_msgs:
                txts = [norm(a.get("text","")) for a in bot_msgs if a.get("text")]
                if txts:
                    return True, txts[-1][:500]
            time.sleep(0.6)
        return False, "no bot reply"
    except Exception as e:
        return False, f"error: {e}"

# ===================== Streamlit UI =====================
st.set_page_config(page_title="Excel抽出ツール", page_icon="🧾", layout="wide")
# タイトルは表示しない（ユーザー要望）

with st.sidebar:
    st.subheader("Excel抽出の必須項目")
    require_lotno = st.checkbox("Lot No.を必須にする", value=True)
    require_exp   = st.checkbox("有効期限を必須にする", value=True)

    st.subheader("Copilot連携テスト（Direct Line）")
    directline_secret = st.text_input("Direct Line シークレット（既定のボット）", type="password")
    test_text = st.text_input("テスト送信メッセージ", value="ping")
    if st.button("Copilot 接続テストを実行"):
        if not directline_secret:
            st.error("Direct Line シークレットを入力してください。")
        else:
            ok, msg = copilot_directline_test(directline_secret, test_message=test_text)
            if ok:
                st.success(f"✅ 連携OK（応答）：{msg}")
            else:
                st.error(f"❌ 連携NG：{msg}")

st.markdown("### 1) 入力ファイル（Excel／複数可、★1回につき最大8ファイルまで）")
xlsx_inputs = st.file_uploader("Excel（シート自動選択：編集用＞最新日付＞その他、確認系は除外）", type=["xlsx"], accept_multiple_files=True)

st.markdown("### 2) 追記先Excel（未指定なら新規作成してDL可）")
out_book = st.file_uploader("既存Excel（“編集用/品名ごと/工程ごと/品名マスタ”を含む想定）", type=["xlsx"])

st.markdown("---")
run = st.button("▶ データ抽出")

# 状態
if "rows_all" not in st.session_state: st.session_state.rows_all=[]
if "problems" not in st.session_state: st.session_state.problems=[]
if "updated_excel_bytes" not in st.session_state: st.session_state.updated_excel_bytes=None

if run:
    st.session_state.rows_all=[]; st.session_state.problems=[]; st.session_state.updated_excel_bytes=None

    # -------- Excel処理のみ --------
    if xlsx_inputs:
        for xf in xlsx_inputs:
            try:
                xbytes=xf.read()
                xls=pd.ExcelFile(io.BytesIO(xbytes))
                target_sheet, reason = choose_target_sheet(xls.sheet_names)
                df=pd.read_excel(io.BytesIO(xbytes), sheet_name=target_sheet, header=None)

                koutei, lot = extract_koutei_lot_from_sheet(df)
                hmap=detect_header(df, scan_rows=60)
                if not hmap:
                    st.session_state.problems.append(f"{xf.name}: ヘッダ検出失敗（{target_sheet} / 理由: {reason}）")
                    continue

                # ファイル名に「返庫」を含む場合、払出数をマイナス符号で取り込む
                base_name = xf.name.rsplit(".", 1)[0]
                qty_sign = -1 if is_henko_from_name(base_name) else 1

                rows, rej = parse_excel_table(
                    df, hmap, koutei or "", lot or "",
                    file_label=base_name,
                    qty_sign=qty_sign,
                    require_lotno=require_lotno,
                    require_exp=require_exp,
                )
                st.session_state.rows_all.extend(rows)
                if not rows:
                    st.session_state.problems.append(
                        f"{xf.name}: 明細0件（{target_sheet} / {reason} / 拒否内訳: {rej}）"
                    )
                else:
                    bad = {k:v for k,v in rej.items() if v>0}
                    if bad:
                        st.info(f"{xf.name}（{target_sheet}）でスキップ: {bad}")
            except Exception as e:
                st.session_state.problems.append(f"{xf.name}: 解析エラー: {e}")

    total=len(st.session_state.rows_all)
    if st.session_state.problems:
        st.warning("一部で問題:\n- " + "\n- ".join(st.session_state.problems))
    if total==0:
        st.error("有効なデータ行を抽出できませんでした。")
    else:
        st.success(f"合計 {total} 行を抽出しました。")

    # -------- “編集用”追記＋レポート再作成 --------
    try:
        base_bytes = out_book.getvalue() if out_book else None
        updated = update_workbook_with_rows(base_bytes, st.session_state.rows_all, sheet_name="編集用")
        st.session_state.updated_excel_bytes = updated
        st.info("『編集用』へ追記し、『品名ごと』『工程ごと』を最新化しました。")
    except Exception as e:
        st.error(f"Excelの更新に失敗: {e}")

# ===================== プレビュー =====================
st.markdown("### 抽出結果（先頭300行）")
if st.session_state.rows_all:
    df_out = pd.DataFrame(st.session_state.rows_all, columns=HEADERS)[:300]
    st.dataframe(df_out, use_container_width=True)

st.markdown("### 更新済みExcelのダウンロード")
if st.session_state.updated_excel_bytes:
    st.download_button(
        "📥 更新済みExcelをダウンロード",
        data=st.session_state.updated_excel_bytes,
        file_name=f"updated_{int(time.time())}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
    )
